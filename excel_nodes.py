import os
import openpyxl
import csv
import torch
import re

class TrucyExcelReader:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "C:\\example.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row": ("INT", {"default": 1, "min": 1, "max": 999999}),
                "column": ("INT", {"default": 1, "min": 1, "max": 999}),
            }
        }

    RETURN_TYPES = ("STRING", "INT", "FLOAT")
    RETURN_NAMES = ("string", "int", "float")
    FUNCTION = "read_cell"
    CATEGORY = "TrucyNodes/Excel"

    @classmethod
    def IS_CHANGED(cls, excel_path, sheet_name, row, column):
        clean_path = excel_path.strip().replace('"', '')
        target_file = clean_path

        # 如果输入的是文件夹，自动寻找第一个 xlsx/csv 文件
        if os.path.isdir(clean_path):
            valid_exts = ['.xlsx', '.csv'] # 彻底废弃老旧的 .xls
            files = [f for f in os.listdir(clean_path) if any(f.lower().endswith(ext) for ext in valid_exts)]
            if files:
                files.sort()
                target_file = os.path.join(clean_path, files[0])
        
        # 监测最终文件的修改时间
        if os.path.isfile(target_file):
            return os.path.getmtime(target_file)
        return float("NaN")

    def read_cell(self, excel_path, sheet_name, row, column):
        clean_path = excel_path.strip().replace('"', '')
        target_file = clean_path

        # --- 智能文件夹探测模式 ---
        if os.path.isdir(clean_path):
            valid_exts = ['.xlsx', '.csv']
            files = [f for f in os.listdir(clean_path) if any(f.lower().endswith(ext) for ext in valid_exts)]
            if files:
                files.sort()
                target_file = os.path.join(clean_path, files[0])
                print(f"[TrucyNodes] Directory provided. Auto-selected file: {target_file}")
            else:
                return ("Error: No .xlsx or .csv files found in directory", 0, 0.0)

        if not os.path.isfile(target_file):
            return (f"Error: File not found at {target_file}", 0, 0.0)

        ext = target_file.lower().split('.')[-1]
        cell_value = None

        try:
            # --- 读取 CSV 逻辑 (无视 Sheet 名称) ---
            if ext == 'csv':
                # 兼容不同编码，优先使用 utf-8-sig 以去除可能存在的 BOM 头
                with open(target_file, 'r', encoding='utf-8-sig', errors='replace') as f:
                    reader = list(csv.reader(f))
                    # 校验行数和列数是否越界 (Excel/CSV 用户习惯从 1 开始)
                    if 1 <= row <= len(reader) and 1 <= column <= len(reader[row-1]):
                        cell_value = reader[row-1][column-1]
                    else:
                        return ("Error: Row or Column out of range in CSV", 0, 0.0)
            
            # --- 读取 Excel (.xlsx) 逻辑 (强校验 Sheet) ---
            elif ext == 'xlsx':
                workbook = openpyxl.load_workbook(target_file, data_only=True, read_only=True)
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    cell_value = sheet.cell(row=row, column=column).value
                else:
                    return (f"Error: Sheet '{sheet_name}' not found", 0, 0.0)
                workbook.close()
            
            # --- 拒绝老旧的 .xls ---
            elif ext == 'xls':
                return ("Error: Legacy .xls format is unsupported. Please save as .xlsx or .csv", 0, 0.0)
            else:
                return ("Error: Unsupported file format", 0, 0.0)

            # --- 空值处理 ---
            if cell_value is None or str(cell_value).strip() == "":
                return ("N/A", 0, 0.0)

            # --- 智能数字提取 ---
            full_str = str(cell_value)
            number_match = re.search(r"[-+]?\d*\.\d+|[-+]?\d+", full_str)

            res_float = 0.0
            res_int = 0

            if number_match:
                num_str = number_match.group() 
                try:
                    res_float = float(num_str)
                    res_int = int(res_float) 
                except:
                    pass

            return (full_str, res_int, res_float)

        except Exception as e:
            return (f"Error: {str(e)}", 0, 0.0)

NODE_CLASS_MAPPINGS = {
    "TrucyExcelReader": TrucyExcelReader
}

NODE_DISPLAY_NAME_MAPPINGS = {
    "TrucyExcelReader": "🚀 Excel Cell Reader (Trucy)"
}